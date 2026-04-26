"""
AI ORCHESTRATOR — Telegram Bot (Polling Mode)
Setiap pesan diproses sebagai asyncio.Task terpisah
agar polling loop tidak terblokir saat AI generate.
"""
import asyncio
import threading
import logging
from typing import Optional
import structlog
import html
import re

log = structlog.get_logger()

_bot_thread: Optional[threading.Thread] = None
_bot_running = False
_bot_loop:   Optional[asyncio.AbstractEventLoop] = None
_chat_tasks: dict[int, asyncio.Task] = {}

# ── Telegram Token Masking ────────────────────────────────────
# Prevents Bot Token from leaking into logs (httpx logs URLs with token)
_current_token: Optional[str] = None

def _mask_token(text: str) -> str:
    """Replace bot token in any string with masked version."""
    if not _current_token or not text:
        return text
    return text.replace(_current_token, _current_token[:8] + "***MASKED***")

class _TelegramLogFilter(logging.Filter):
    """Filter that masks the Telegram Bot Token from httpx HTTP log lines."""
    def filter(self, record: logging.LogRecord) -> bool:
        if _current_token and isinstance(record.msg, str):
            record.msg = _mask_token(record.msg)
        if _current_token and record.args:
            try:
                record.args = tuple(
                    _mask_token(str(a)) if isinstance(a, str) else a
                    for a in record.args
                )
            except Exception:
                pass
        return True

# Apply filter to httpx logger so token never appears in logs
_httpx_logger = logging.getLogger("httpx")
_httpx_logger.addFilter(_TelegramLogFilter())


async def _resolve_web_user(chat_id: int, fallback_user_id: str) -> str:
    """Mencari Web User ID berdasarkan telegram_chat_id. Fallback ke ID Telegram jika tidak ada."""
    try:
        from db.database import AsyncSessionLocal
        from db.models import User
        from sqlmodel import select
        async with AsyncSessionLocal() as db:
            result = await db.execute(select(User).where(User.telegram_chat_id == str(chat_id)))
            user = result.scalar_one_or_none()
            if user:
                return user.id
    except Exception:
        pass
    return fallback_user_id


# ── Helper kirim pesan ────────────────────────────────────────
async def _send(token: str, chat_id: int, text: str, include_drive_btn: bool = False, max_retries: int = 3):
    """Kirim pesan Markdown, auto-split > 4000 char dengan retry mechanism."""
    import httpx
    import asyncio
    
    chunks = [text[i:i+4000] for i in range(0, max(len(text), 1), 4000)]
    
    for retry in range(max_retries):
        try:
            async with httpx.AsyncClient(timeout=30) as c:  # Increased timeout
                for idx, chunk in enumerate(chunks):
                    payload = {"chat_id": chat_id, "text": chunk, "parse_mode": "HTML"}
                    if include_drive_btn and idx == len(chunks) - 1:
                        payload["reply_markup"] = {
                            "inline_keyboard": [
                                [
                                    {"text": "🔊 Dengar (Voice)", "callback_data": "voice_reply"},
                                    {"text": "⬇️ Download File", "callback_data": "download_options"}
                                ]
                            ]
                        }
                    
                    # Retry individual chunk if needed
                    chunk_retry = 0
                    while chunk_retry < 2:
                        try:
                            await c.post(
                                "https://api.telegram.org/bot" + token + "/sendMessage",
                                json=payload,
                            )
                            break  # Success, move to next chunk
                        except Exception as chunk_ex:
                            chunk_retry += 1
                            if chunk_retry >= 2:
                                raise chunk_ex
                            await asyncio.sleep(1)  # Wait before retry
                
                # All chunks sent successfully
                return
                
        except Exception as ex:
            log.warning("Telegram send error (retry {}/{})".format(retry + 1, max_retries), error=str(ex)[:80])
            if retry < max_retries - 1:
                await asyncio.sleep(2 ** retry)  # Exponential backoff
            else:
                # Final retry failed, log and continue
                log.error("Telegram send failed after all retries", chat_id=chat_id, error=str(ex)[:100])


async def _send_voice(token: str, chat_id: int, text: str):
    """Generate TTS and send as voice note."""
    import httpx, io, edge_tts
    try:
        communicate = edge_tts.Communicate(text, "id-ID-ArdiNeural")
        audio_data = io.BytesIO()
        async for chunk in communicate.stream():
            if chunk["type"] == "audio":
                audio_data.write(chunk["data"])
        audio_data.seek(0)
        
        async with httpx.AsyncClient(timeout=30) as c:
            await c.post(
                f"https://api.telegram.org/bot{token}/sendVoice",
                data={'chat_id': str(chat_id)},
                files={'voice': ('voice.mp3', audio_data.getvalue())}
            )
    except Exception as e:
        log.error("Telegram send_voice error", error=str(e))

async def _send_typing(token: str, chat_id: int):
    import httpx
    try:
        async with httpx.AsyncClient(timeout=5) as c:
            await c.post(
                "https://api.telegram.org/bot" + token + "/sendChatAction",
                json={"chat_id": chat_id, "action": "typing"},
            )
    except Exception:
        pass

async def _edit_status(token: str, chat_id: int, message_id: int, text: str):
    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            await c.post(
                f"https://api.telegram.org/bot{token}/editMessageText",
                json={
                    "chat_id": chat_id,
                    "message_id": message_id,
                    "text": text,
                    "parse_mode": "HTML",
                    "reply_markup": {
                        "inline_keyboard": [[{"text": "❌ Berhenti", "callback_data": "cancel_process"}]]
                    }
                }
            )
    except Exception:
        pass

def _escape(text: str) -> str:
    """Helper to escape HTML for Telegram safety."""
    if not text: return ""
    return html.escape(text)

def _format_for_tg(text: str) -> str:
    """Convert common AI markdown to Telegram-safe HTML."""
    if not text: return ""
    # 1. Escape basic HTML tags
    t = _escape(text)
    # 2. Convert bold ** to <b>
    t = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', t)
    # 3. Convert single * to <i> (if not inside <b>)
    t = re.sub(r'\*(.*?)\*', r'<i>\1</i>', t)
    # 4. Convert inline code ` to <code>
    t = re.sub(r'`\s*(.*?)\s*`', r'<code>\1</code>', t)
    # 5. Convert code blocks ``` to <pre>
    def code_block_sub(match):
        content = match.group(2).strip()
        return f'<pre>{content}</pre>'
    t = re.sub(r'```(\w+)?\n?(.*?)```', code_block_sub, t, flags=re.DOTALL)
    return t


# ── Handle satu pesan (asyncio.Task terpisah) ─────────────────
async def _handle_message(chat_id: int, user_id: str, text: str,
                           from_name: str, token: str):
    try:
        # Resolve to Web User ID if linked
        user_id = await _resolve_web_user(chat_id, user_id)

        # /start
        if text == "/start":
            await _send(token, chat_id,
                "Halo " + from_name + "! Saya *AI ORCHESTRATOR*, AI Orchestrator kamu.\n\n"
                "Ketik pertanyaan apa saja dan saya akan menjawab!\n\n"
                "Perintah:\n"
                "/start - Mulai\n"
                "/help - Bantuan\n"
                "/clear - Hapus konteks chat\n"
                "/model - Lihat model aktif"
            )
            return

        # /help
        if text == "/help":
            await _send(token, chat_id,
                "*AI ORCHESTRATOR Bot*\n\n"
                "Kirim pesan teks biasa untuk chat dengan AI.\n\n"
                "Fitur:\n"
                "- Chat dengan AI (auto-pilih model terbaik)\n"
                "- Cari di knowledge base (RAG)\n"
                "- Ingat konteks percakapan\n\n"
                "Perintah:\n"
                "/start - Mulai bot\n"
                "/help - Tampilkan bantuan ini\n"
                "/clear - Reset konteks chat\n"
                "/model - Model yang digunakan"
            )
            return

        # /clear
        if text == "/clear":
            from memory.manager import memory_manager
            await memory_manager.save_short_term("tg_" + str(chat_id), [])
            await _send(token, chat_id, "Konteks chat direset!")
            return

        # /model
        if text == "/model":
            from core.model_manager import model_manager
            default_model = model_manager.get_default_model()
            models_list = ", ".join(list(model_manager.available_models.keys())[:5])
            await _send(token, chat_id, f"Model default: `{default_model}`\nModel tersedia: `{models_list}`")
            return

        # Pesan biasa → proses AI
        await _send_typing(token, chat_id)

        from core.orchestrator import orchestrator
        from core.model_manager import model_manager
        from core.smart_router import smart_router
        from memory.manager import memory_manager
        from rag.engine import rag_engine
        from db.database import AsyncSessionLocal
        from db.models import Message, ChatSession
        from sqlmodel import select

        route  = smart_router.route(text)
        # Check if default model exists before using it
        if "sumopod/seed-2-0-pro" in model_manager.available_models:
            model = "sumopod/seed-2-0-pro"
        else:
            # Fallback to available model or default
            available_models = list(model_manager.available_models.keys())
            model = model_manager.get_default_model() if available_models else available_models[0]
        system = await memory_manager.build_system_prompt(user_id, "tg_" + str(chat_id))

        # Telegram context
        system += "\n\n[PENTING: SUMBER TELEGRAM]\nAnda sedang berkomunikasi melalui bot Telegram. Jika pengguna meminta untuk membuat file, script, atau aplikasi, JANGAN menyimpannya di root direktori AI Orchestrator. Selalu simpan file-file yang dihasilkan ke dalam folder terpisah (misalnya /home/bamuskal/Documents/ai-super/data/generated_apps/). Selalu buat folder tujuan menggunakan command line dengan format absolut dan gunakan path absolut tersebut di semua perintah file/folder selanjutnya (contoh: `mkdir -p /absolute/path/nama-app` lalu `cd /absolute/path/nama-app`). DILARANG KERAS menggunakan perintah `cd` dengan path relatif karena sesi eksekusi terminal tidak persisten antar-langkah."

        # RAG context
        try:
            rag_results = await rag_engine.query(text, user_id=user_id)
            if rag_results:
                system += "\n\n" + rag_engine.build_context(rag_results)
        except Exception:
            pass

        # Riwayat chat (load dari Redis, fallback ke DB)
        history  = await memory_manager.get_context("tg_" + str(chat_id), user_id)
        messages = [{"role": "system", "content": system}]
        messages += history
        messages.append({"role": "user", "content": text})

        # Generate using Orchestrator for full monitoring and robustness
        full_response = ""
        session_id = "tg_" + str(chat_id)
        status_msg_id = None
        
        # Ensure session exists in DB for monitoring
        async with AsyncSessionLocal() as db:
            sess = await db.get(ChatSession, session_id)
            if not sess:
                sess = ChatSession(id=session_id, user_id=user_id, title=text[:50], platform="telegram")
                db.add(sess)
                await db.commit()
            
            # Save User Message to DB
            user_msg = Message(session_id=session_id, user_id=user_id, role="user", content=text)
            db.add(user_msg)
            await db.commit()

        # Periodic typing status task
        typing_active = True
        async def keep_typing():
            while typing_active:
                await _send_typing(token, chat_id)
                await asyncio.sleep(5)
        
        typing_task = asyncio.create_task(keep_typing())
        
        try:
            _chat_tasks[chat_id] = asyncio.current_task()
            # Add timeout wrapper for orchestrator processing
            chunk_count = 0
            async def process_with_timeout():
                nonlocal chunk_count, full_response
                try:
                    async for event in orchestrator.process(
                        message=text,
                        user_id=user_id,
                        session_id=session_id,
                        system_prompt=system,
                        history=history,
                        include_tool_logs=False,
                        emit_thinking=False,
                        auto_execute=True,
                    ):
                        if event.type == "chunk":
                            full_response += event.content
                            chunk_count += 1
                            log.debug("Telegram chunk received", chat_id=chat_id, chunk_num=chunk_count, content_preview=event.content[:50])
                        elif event.type == "process":
                            # Dynamic status updates
                            action = event.data.get("action", "Thinking")
                            detail = event.data.get("detail", "")
                            status_text = f"⚙️ <b>{action}...</b>"
                            if detail:
                                status_text += f"\n<i>{_escape(detail)}</i>"
                            
                            if status_msg_id is None:
                                try:
                                    import httpx
                                    async with httpx.AsyncClient(timeout=10) as c:
                                        r = await c.post(f"https://api.telegram.org/bot{token}/sendMessage", 
                                                    json={
                                                        "chat_id": chat_id, 
                                                        "text": status_text, 
                                                        "parse_mode": "HTML",
                                                        "reply_markup": {
                                                            "inline_keyboard": [[{"text": "❌ Berhenti", "callback_data": "cancel_process"}]]
                                                        }
                                                    })
                                        status_msg_id = r.json().get("result", {}).get("message_id")
                                except Exception:
                                    pass
                            else:
                                await _edit_status(token, chat_id, status_msg_id, status_text)

                        elif event.type == "error":
                            full_response = f"⚠️ <b>Error:</b> {_escape(event.content)}"
                            log.error("Telegram orchestrator error", chat_id=chat_id, error=event.content)
                            break
                        elif event.type == "done":
                            # Clear status message if it exists
                            if status_msg_id:
                                try:
                                    import httpx
                                    async with httpx.AsyncClient(timeout=5) as c:
                                        await c.post(f"https://api.telegram.org/bot{token}/deleteMessage", 
                                                   json={"chat_id": chat_id, "message_id": status_msg_id})
                                except Exception:
                                    pass
                            log.info("Telegram orchestrator done", chat_id=chat_id, total_chunks=chunk_count, response_length=len(full_response))
                        elif event.type == "pending_confirmation":
                            # Should not happen with auto_execute=True, but handle gracefully
                            full_response = f"⚠️ Perintah ini memerlukan konfirmasi: <code>{_escape(event.data.get('command', text))}</code>\n\nSilakan gunakan web dashboard untuk menjalankan perintah ini."
                            break
                except asyncio.TimeoutError:
                    full_response = "⏱️ Maaf, response terlalu lama. Silakan coba lagi dengan pertanyaan yang lebih singkat."
                except Exception as e:
                    log.error("Orchestrator processing error", error=str(e), chat_id=chat_id)
                    full_response = f"⚠️ Terjadi kesalahan saat memproses: {str(e)[:100]}"
            
            # Set timeout to 300 seconds (5 minutes) for complex tasks like app development
            await asyncio.wait_for(process_with_timeout(), timeout=300.0)
            
            log.info("Telegram processing complete", chat_id=chat_id, total_chunks=chunk_count, response_length=len(full_response), has_content=bool(full_response.strip()))
            
        except asyncio.TimeoutError:
            full_response = "⏱️ Maaf, request terlalu lama diproses. Silakan coba lagi."
            log.warning("Telegram message processing timeout", chat_id=chat_id)
        except Exception as e:
            log.error("Telegram message processing failed", error=str(e), chat_id=chat_id)
            full_response = f"⚠️ Maaf, terjadi kesalahan: {str(e)[:80]}..."
        finally:
            if _chat_tasks.get(chat_id) == asyncio.current_task():
                del _chat_tasks[chat_id]
            typing_active = False
            try:
                typing_task.cancel()
            except Exception:
                pass

        if not full_response.strip():
            log.warning("Telegram empty response", chat_id=chat_id, text=text[:50])
            full_response = "⚠️ AI tidak memberikan respons. Ini mungkin karena model sedang sibuk atau pertanyaan terlalu kompleks."

        await _send(token, chat_id, _format_for_tg(full_response), include_drive_btn=True)

        # Save Assistant Message to DB
        async with AsyncSessionLocal() as db:
            ai_msg = Message(session_id=session_id, user_id=user_id, role="assistant", content=full_response, model=model)
            db.add(ai_msg)
            
            # Update session timestamp
            sess = await db.get(ChatSession, session_id)
            if sess:
                from datetime import datetime
                sess.updated_at = datetime.utcnow()
                db.add(sess)
            await db.commit()

        # Simpan ke Redis (Memory)
        await memory_manager.save_chat_to_redis(session_id, "user", text)
        await memory_manager.save_chat_to_redis(session_id, "assistant", full_response)
        log.info("Telegram reply sent (Orchestrated)", chat_id=chat_id, model=model, chars=len(full_response))

    except asyncio.CancelledError:
        log.info("Telegram task cancelled by user", chat_id=chat_id)
        return
    except Exception as e:
            log.error("Telegram handle_message error", error=str(e), chat_id=chat_id)
            err = str(e).lower()
            try:
                if "cancelled" in err:
                    return
                if "timeout" in err or "timed out" in err:
                    msg = "⏱️ Maaf, request terlalu lama. VPS mungkin sedang load tinggi. Coba kirim lagi ya!"
                elif "api key" in err or "unauthorized" in err or "invalid" in err:
                    msg = "🔑 Model AI tidak bisa diakses. Pastikan API key sudah diset di menu Integrasi."
                elif "redis" in err or "connection" in err:
                    msg = "🔄 Sedang menginisialisasi koneksi VPS. Coba kirim pesan lagi."
                elif "memory" in err or "resource" in err:
                    msg = "💾 VPS resource sedang tinggi. Silakan coba beberapa saat lagi."
                else:
                    msg = "⚠️ VPS sedang ada gangguan sementara. Silakan kirim ulang pesan kamu."
                await _send(token, chat_id, msg)
            except Exception:
                pass

# ── Handle pesan FOTO ────────────────────────────────────────
async def _handle_photo(chat_id: int, user_id: str, photo_list: list, caption: str,
                        from_name: str, token: str):
    """Download foto terbesar dari Telegram, kirim ke model vision."""
    try:
        # Resolve to Web User ID if linked
        user_id = await _resolve_web_user(chat_id, user_id)
        
        import httpx, base64
        # Foto terbesar = index terakhir
        best_photo = photo_list[-1]
        file_id = best_photo["file_id"]

        await _send_typing(token, chat_id)

        # 1. Get file path
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.get(f"https://api.telegram.org/bot{token}/getFile",
                            params={"file_id": file_id})
            file_path = r.json()["result"]["file_path"]

        # 2. Download file
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(f"https://api.telegram.org/file/bot{token}/{file_path}")
            image_bytes = r.content

        # 3. Encode ke base64
        image_b64 = base64.b64encode(image_bytes).decode("utf-8")
        ext = file_path.split(".")[-1].lower() if "." in file_path else "jpg"
        mime = "image/jpeg" if ext in ("jpg", "jpeg") else f"image/{ext}"

        # 4. Kirim ke vision model
        from core.model_manager import model_manager
        from memory.manager import memory_manager

        system = await memory_manager.build_system_prompt(user_id, "tg_" + str(chat_id))
        user_question = caption or "Deskripsikan gambar ini secara detail dalam Bahasa Indonesia."
        history = await memory_manager.get_context("tg_" + str(chat_id), user_id)

        full_response = await model_manager.chat_with_image(
            image_b64=image_b64,
            mime_type=mime,
            text_prompt=user_question,
            system_prompt=system,
            history=history,
        )

        if not full_response.strip():
            full_response = "Maaf, saya tidak bisa menganalisis gambar ini. Coba kirim ulang ya!"

        await _send(token, chat_id, full_response, include_drive_btn=True)

        # Simpan ke memory
        await memory_manager.save_chat_to_redis("tg_" + str(chat_id), "user", f"[Foto] {user_question}")
        await memory_manager.save_chat_to_redis("tg_" + str(chat_id), "assistant", full_response)
        log.info("Telegram photo handled", chat_id=chat_id, chars=len(full_response))

    except Exception as e:
        log.error("Telegram handle_photo error", error=str(e))
        try:
            await _send(token, chat_id, "Maaf, gagal memproses gambar. Coba kirim ulang ya!")
        except Exception:
            pass


# ── Handle pesan SUARA/VOICE ──────────────────────────────────
async def _handle_voice(chat_id: int, user_id: str, voice_obj: dict,
                        from_name: str, token: str):
    """Download voice note dari Telegram, transkrip via Whisper, lanjutkan sebagai teks."""
    try:
        # Resolve to Web User ID if linked
        user_id = await _resolve_web_user(chat_id, user_id)
        
        import httpx
        file_id = voice_obj["file_id"]

        await _send(token, chat_id, "🎙️ _Mentranskripsi pesan suara..._")
        await _send_typing(token, chat_id)

        # 1. Get file path
        async with httpx.AsyncClient(timeout=15) as c:
            r = await c.get(f"https://api.telegram.org/bot{token}/getFile",
                            params={"file_id": file_id})
            file_path = r.json()["result"]["file_path"]

        # 2. Download audio
        async with httpx.AsyncClient(timeout=30) as c:
            r = await c.get(f"https://api.telegram.org/file/bot{token}/{file_path}")
            audio_bytes = r.content

        # 3. Transkrip via model_manager
        from core.model_manager import model_manager
        transcript = await model_manager.transcribe_audio(
            audio_bytes=audio_bytes,
            filename=file_path.split("/")[-1] or "voice.ogg",
        )

        if not transcript or not transcript.strip():
            await _send(token, chat_id, "Maaf, tidak bisa mentranskrip suara. Kirim pesan teks ya!")
            return

        await _send(token, chat_id, f"🎙️ _Transkripsi:_ \"{transcript}\"")

        # 4. Lanjutkan sebagai pesan teks biasa
        await _handle_message(chat_id, user_id, transcript, from_name, token)

    except Exception as e:
        log.error("Telegram handle_voice error", error=str(e))
        try:
            await _send(token, chat_id, "Maaf, gagal memproses pesan suara. Silakan kirim ulang.")
        except Exception:
            pass

# ── Handle Callback Query ─────────────────────────────────────
async def _handle_callback_query(callback_query: dict, token: str):
    import httpx
    from memory.manager import memory_manager
    import json
    
    chat_id = callback_query.get("message", {}).get("chat", {}).get("id")
    msg_id = callback_query.get("message", {}).get("message_id")
    data = callback_query.get("data", "")
    callback_id = callback_query.get("id")
    user_id = str(callback_query.get("from", {}).get("id", "unknown"))
    
    async def _answer(text=""):
        async with httpx.AsyncClient() as c:
            await c.post(f"https://api.telegram.org/bot{token}/answerCallbackQuery", json={"callback_query_id": callback_id, "text": text})
            
    async def _edit_markup(markup):
        async with httpx.AsyncClient() as c:
            await c.post(f"https://api.telegram.org/bot{token}/editMessageReplyMarkup", json={"chat_id": chat_id, "message_id": msg_id, "reply_markup": markup})
            
    async def _send_text(text):
        await _send(token, chat_id, text)

    try:
        if data == "cancel_process":
            await _answer()
            task = _chat_tasks.get(chat_id)
            if task and not task.done():
                task.cancel()
                await _send_text("❌ Proses telah dibatalkan.")
                try:
                    import httpx
                    async with httpx.AsyncClient(timeout=5) as c:
                        await c.post(f"https://api.telegram.org/bot{token}/deleteMessage", 
                                   json={"chat_id": chat_id, "message_id": msg_id})
                except Exception:
                    pass
            return

        if data == "voice_reply":
            await _answer("Menyiapkan suara...")
            history = await memory_manager.get_context("tg_" + str(chat_id), user_id)
            if history:
                text = history[-1]['content']
                await _send_voice(token, chat_id, text)
            return

        if data == "download_options":
            await _answer("Pilih Format")
            markup = {"inline_keyboard": [
                [{"text": "📄 PDF", "callback_data": "download_format_pdf"}, {"text": "📝 Word", "callback_data": "download_format_docx"}],
                [{"text": "📊 Excel", "callback_data": "download_format_xlsx"}, {"text": "🗄 TXT", "callback_data": "download_format_txt"}]
            ]}
            await _edit_markup(markup)
            
        elif data.startswith("download_format_"):
            fmt = data.split("_")[-1]
            await memory_manager.save_short_term(f"tg_fmt_{chat_id}", fmt) # store format
            
            # Fetch content
            history = await memory_manager.get_context("tg_" + str(chat_id), user_id)
            if not history:
                await _send_text("❌ Gagal menyusun dokumen. Konten sudah kadaluarsa.")
                return
            last_ai = history[-1]['content']
            
            await _process_and_send_file(token, chat_id, last_ai, fmt, action="download")

    except Exception as e:
        log.error("Telegram callback error", error=str(e))
        await _answer()

async def _process_and_send_file(token, chat_id, last_ai, fmt, action="download"):
    import httpx
    import io, re
    
    filename = f"AI_Generated_{chat_id}.{fmt}"
    file_bytes = b""

            
    try:
        if fmt in ["txt", "csv", "md"]:
            file_bytes = last_ai.encode("utf-8")
        elif fmt == "docx":
            from docx import Document
            doc = Document()
            for line in last_ai.split("\n"): doc.add_paragraph(line)
            buf = io.BytesIO()
            doc.save(buf)
            file_bytes = buf.getvalue()
        elif fmt == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            lines = last_ai.split('\n')
            table_rows = []
            regular_text = []
            for line in lines:
                line = line.strip()
                if line.startswith('|') and line.endswith('|'):
                    if re.match(r'^\|[-\s\|:]+\|$', line): 
                        continue
                    cells = [c.strip() for c in line.split('|')[1:-1]]
                    if cells: table_rows.append(cells)
                else:
                    if line: regular_text.append(line)
            current_row = 1
            if regular_text:
                context = "\n".join(regular_text)
                ws.cell(row=current_row, column=1, value="Note / Context:").font = Font(bold=True)
                ws.merge_cells(start_row=current_row+1, start_column=1, end_row=current_row+1, end_column=10)
                ctx_cell = ws.cell(row=current_row+1, column=1, value=context)
                ctx_cell.alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 3
            if table_rows:
                header_font = Font(name="Calibri", bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                cell_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'), top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
                for c_idx, val in enumerate(table_rows[0], 1):
                    cell = ws.cell(row=current_row, column=c_idx, value=val)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = cell_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                for r_idx, row_data in enumerate(table_rows[1:], 1):
                    for c_idx, val in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row + r_idx, column=c_idx, value=val)
                        cell.border = cell_border
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
            buf = io.BytesIO()
            wb.save(buf)
            file_bytes = buf.getvalue()
        elif fmt == "pdf":
            from fpdf import FPDF, HTMLMixin
            import markdown
            class TGPDF(FPDF, HTMLMixin):
                pass
            pdf = TGPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", size=10)
            html_content = markdown.markdown((last_ai or "").strip(), extensions=['tables', 'fenced_code'])
            html_clean = html_content.encode('latin-1', 'replace').decode('latin-1')
            pdf.write_html(html_clean)
            buf = io.BytesIO()
            pdf.output(buf)
            file_bytes = buf.getvalue()
    except Exception as exc:
        log.error("File generation failed", error=str(exc))
        async with httpx.AsyncClient() as c:
            await c.post(f"https://api.telegram.org/bot{token}/sendMessage", json={"chat_id": chat_id, "text": "❌ Terjadi distorsi saat menyusun file."})
        return
        
    # Download logic
    try:
        async with httpx.AsyncClient(timeout=60) as c:
            await c.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={'chat_id': str(chat_id)},
                files={'document': (filename, file_bytes)}
            )
    except Exception as e:
        log.error("Telegram sendDocument error", error=str(e)[:80])
        async with httpx.AsyncClient() as c:
            await c.post(f"https://api.telegram.org/bot{token}/sendMessage", json={"chat_id": chat_id, "text": "❌ Gagal mengirim dokumen ke chat."})


# ── Callback task selesai ─────────────────────────────────────
def _task_done(task: asyncio.Task, active_tasks: set):
    active_tasks.discard(task)
    if task.cancelled():
        return
    try:
        exc = task.exception()
        if exc:
            log.error("Telegram task error", error=str(exc))
    except Exception:
        pass


# ── Polling loop utama ────────────────────────────────────────
async def _polling_loop(token: str):
    global _bot_running
    import httpx

    offset       = 0
    active_tasks: set = set()

    # Tunggu service siap
    await asyncio.sleep(2)
    log.info("Telegram polling started")

    # Flush pending updates
    try:
        async with httpx.AsyncClient(timeout=10) as c:
            r = await c.get(
                "https://api.telegram.org/bot" + token + "/getUpdates",
                params={"offset": -1, "limit": 1},
            )
            data = r.json()
            if data.get("ok") and data.get("result"):
                offset = data["result"][-1]["update_id"] + 1
                log.info("Flushed pending updates", next_offset=offset)
    except Exception as e:
        log.warning("Flush updates failed", error=str(e))

    consecutive_errors = 0
    max_consecutive_errors = 5
    
    while _bot_running:
        try:
            async with httpx.AsyncClient(timeout=45) as c:  # Increased timeout
                resp = await c.get(
                    "https://api.telegram.org/bot" + token + "/getUpdates",
                    params={
                        "offset":          offset,
                        "timeout":         25,  # Reduced long polling timeout
                        "allowed_updates": ["message", "callback_query"],
                    },
                )
            
            # Reset error counter on successful request
            consecutive_errors = 0

            if resp.status_code == 409:
                log.warning("Telegram 409 Conflict, retry 10s")
                await asyncio.sleep(10)
                continue

            if resp.status_code != 200:
                log.warning("Telegram non-200", status=resp.status_code)
                await asyncio.sleep(3)
                continue

            data = resp.json()
            if not data.get("ok"):
                log.warning("Telegram not ok", desc=data.get("description"))
                await asyncio.sleep(3)
                continue

            for update in data.get("result", []):
                offset    = update["update_id"] + 1
                # Process Callbacks
                if "callback_query" in update:
                    task = asyncio.create_task(
                        _handle_callback_query(update["callback_query"], token),
                        name="tg-cb-" + str(update["callback_query"]["id"]),
                    )
                    active_tasks.add(task)
                    task.add_done_callback(lambda t: _task_done(t, active_tasks))
                    continue

                message   = update.get("message", {})
                chat_id   = message.get("chat", {}).get("id")
                text      = (message.get("text") or "").strip()
                user_id   = str(message.get("from", {}).get("id", "unknown"))
                from_name = message.get("from", {}).get("first_name", "User")

                photo     = message.get("photo")
                voice     = message.get("voice")
                audio     = message.get("audio")
                
                # Handle gambar
                if photo:
                    caption = (message.get("caption") or "").strip()
                    task = asyncio.create_task(
                        _handle_photo(chat_id, user_id, photo, caption, from_name, token),
                        name="tg-photo-" + str(chat_id),
                    )
                    active_tasks.add(task)
                    task.add_done_callback(lambda t: _task_done(t, active_tasks))
                    continue

                # Handle pesan suara
                if voice or audio:
                    voice_obj = voice or audio
                    task = asyncio.create_task(
                        _handle_voice(chat_id, user_id, voice_obj, from_name, token),
                        name="tg-voice-" + str(chat_id),
                    )
                    active_tasks.add(task)
                    task.add_done_callback(lambda t: _task_done(t, active_tasks))
                    continue

                if not chat_id or not text:
                    continue

                log.info("Message", from_name=from_name, text=text[:50])

                task = asyncio.create_task(
                    _handle_message(chat_id, user_id, text, from_name, token),
                    name="tg-" + str(chat_id),
                )
                active_tasks.add(task)
                task.add_done_callback(lambda t: _task_done(t, active_tasks))

        except asyncio.CancelledError:
            break
        except httpx.ReadTimeout:
            consecutive_errors += 1
            if consecutive_errors >= max_consecutive_errors:
                log.error("Too many consecutive timeouts, restarting polling")
                await asyncio.sleep(10)
                consecutive_errors = 0
            continue
        except Exception as e:
            consecutive_errors += 1
            if _bot_running:
                log.error("Polling error ({}/{})".format(consecutive_errors, max_consecutive_errors), error=str(e))
                if consecutive_errors >= max_consecutive_errors:
                    log.error("Too many consecutive errors, waiting longer")
                    await asyncio.sleep(15)
                    consecutive_errors = 0
                else:
                    await asyncio.sleep(2 ** min(consecutive_errors, 5))  # Exponential backoff

    if active_tasks:
        await asyncio.gather(*active_tasks, return_exceptions=True)

    log.info("Telegram polling stopped")


# ── Thread runner ─────────────────────────────────────────────
def _run_in_thread(token: str):
    global _bot_loop
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    _bot_loop = loop
    try:
        loop.run_until_complete(_polling_loop(token))
    except Exception as e:
        log.error("Polling thread fatal", error=str(e))
    finally:
        loop.close()
        _bot_loop = None


import fcntl
import os

_lock_fd = None

# ── Public API ────────────────────────────────────────────────
def start_polling(token: str) -> bool:
    global _bot_thread, _bot_running, _lock_fd, _current_token
    if _bot_running:
        return False

    # Activate token masking before any network activity
    _current_token = token

    # Ensure only one worker runs the polling using a file lock
    try:
        lock_file = "/tmp/telegram_bot_polling.lock"
        _lock_fd = os.open(lock_file, os.O_CREAT | os.O_RDWR)
        fcntl.flock(_lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        log.info("Telegram polling already running in another worker process")
        return False
    except Exception as e:
        log.warning("Could not acquire lock for telegram polling", error=str(e))

    _bot_running = True
    _bot_thread  = threading.Thread(
        target=_run_in_thread,
        args=(token,),
        daemon=True,
        name="telegram-polling",
    )
    _bot_thread.start()
    log.info("Telegram polling thread started")
    return True


def stop_polling():
    global _bot_running, _bot_thread, _bot_loop, _lock_fd, _current_token
    _bot_running = False
    if _bot_loop and _bot_loop.is_running():
        _bot_loop.call_soon_threadsafe(_bot_loop.stop)
    if _bot_thread and _bot_thread.is_alive():
        _bot_thread.join(timeout=8)
    _bot_thread = None

    if _lock_fd is not None:
        try:
            fcntl.flock(_lock_fd, fcntl.LOCK_UN)
            os.close(_lock_fd)
        except Exception:
            pass
        _lock_fd = None

    _current_token = None  # Clear token masking
    log.info("Telegram polling stopped")


def is_running() -> bool:
    return (
        _bot_running
        and _bot_thread is not None
        and _bot_thread.is_alive()
    )


# ── Polling Watchdog ──────────────────────────────────────────
# Monitors the polling thread and auto-restarts it if it dies unexpectedly.
_watchdog_thread: Optional[threading.Thread] = None
_watchdog_running = False


def _watchdog_loop(token: str, check_interval: int = 60):
    """Background loop that checks polling health and restarts if dead."""
    global _bot_running, _bot_thread
    import time
    log.info("Telegram polling watchdog started", interval=check_interval)

    while _watchdog_running:
        time.sleep(check_interval)

        if not _watchdog_running:
            break

        # Only act if we SHOULD be running but the thread is dead
        if _bot_running and (_bot_thread is None or not _bot_thread.is_alive()):
            log.warning("Telegram polling thread died unexpectedly — restarting")
            try:
                # Release stale lock before restarting
                global _lock_fd
                if _lock_fd is not None:
                    try:
                        import fcntl as _fcntl
                        _fcntl.flock(_lock_fd, _fcntl.LOCK_UN)
                        os.close(_lock_fd)
                    except Exception:
                        pass
                    _lock_fd = None

                _bot_running = False  # Reset so start_polling can proceed
                start_polling(token)
                log.info("Telegram polling restarted by watchdog")
            except Exception as e:
                log.error("Watchdog failed to restart polling", error=str(e))

    log.info("Telegram polling watchdog stopped")


def start_watchdog(token: str, check_interval: int = 60):
    """Start the watchdog thread. Call after start_polling()."""
    global _watchdog_thread, _watchdog_running
    if _watchdog_running and _watchdog_thread and _watchdog_thread.is_alive():
        return  # Already running
    _watchdog_running = True
    _watchdog_thread = threading.Thread(
        target=_watchdog_loop,
        args=(token, check_interval),
        daemon=True,
        name="telegram-watchdog",
    )
    _watchdog_thread.start()


def stop_watchdog():
    """Stop the watchdog thread."""
    global _watchdog_running
    _watchdog_running = False
